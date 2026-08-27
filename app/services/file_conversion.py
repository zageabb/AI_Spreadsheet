"""Workbook file conversion helpers for Excel and CSV formats.

This module intentionally keeps conversion logic separate from the spreadsheet
engine and storage adapters.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook as OpenPyxlWorkbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, range_boundaries

from app.models.workbook import Workbook
from app.services.ooxml_preservation import (
    OOXML_METADATA_KEY, OOXMLPreservationError, OOXMLPreservationLayer,
)


_INVALID_SHEET_CHARS = set('[]:*?/\\')


class WorkbookConversionError(Exception):
    """Raised when workbook conversion fails."""


class WorkbookFileConverter:
    """Import/export workbook files using practical compatibility rules."""

    def __init__(self, ooxml_layer: OOXMLPreservationLayer | None = None) -> None:
        self.ooxml_layer = ooxml_layer or OOXMLPreservationLayer()

    def import_xlsx(self, path: str) -> Workbook:
        """Import an `.xlsx` file into the app workbook model."""
        file_path = Path(path)
        try:
            formulas_wb = load_workbook(filename=file_path, data_only=False)
            values_wb = load_workbook(filename=file_path, data_only=True)
        except Exception as exc:  # pragma: no cover - defensive wrapper
            raise WorkbookConversionError(f"Failed to read Excel file: {path}") from exc

        workbook = Workbook(name=file_path.stem or "Imported Workbook")
        workbook.sheets = []
        workbook.metadata["defined_names"] = [
            {
                "name": name,
                "refers_to": defined.attr_text,
                "scope": None,
            }
            for name, defined in formulas_wb.defined_names.items()
            if defined.attr_text and not defined.hidden
        ]
        for source_sheet in formulas_wb.worksheets:
            workbook.metadata["defined_names"].extend({
                "name": name, "refers_to": defined.attr_text, "scope": source_sheet.title,
            } for name,defined in source_sheet.defined_names.items() if defined.attr_text and not defined.hidden)
        try:
            workbook.metadata[OOXML_METADATA_KEY] = self.ooxml_layer.capture(file_path)
        except OOXMLPreservationError as exc:
            raise WorkbookConversionError(f"Failed to preserve Excel OOXML package: {exc}") from exc

        for sheet_index, source_sheet in enumerate(formulas_wb.worksheets):
            imported_sheet = workbook.add_sheet(source_sheet.title)
            imported_sheet.metadata["ooxml_original_title"] = source_sheet.title
            imported_sheet.metadata["freeze_panes"] = str(source_sheet.freeze_panes) if source_sheet.freeze_panes else None
            imported_sheet.metadata["merged_ranges"] = [str(item) for item in source_sheet.merged_cells.ranges]
            imported_sheet.metadata["auto_filter"] = source_sheet.auto_filter.ref
            imported_sheet.metadata["column_widths"] = {
                key: value.width for key, value in source_sheet.column_dimensions.items() if value.width is not None
            }
            imported_sheet.metadata["row_heights"] = {
                str(key): value.height for key, value in source_sheet.row_dimensions.items() if value.height is not None
            }
            imported_sheet.metadata["tables"] = [
                {
                    "name": table.name,
                    "display_name": table.displayName,
                    "ref": table.ref,
                    "columns": [column.name for column in table.tableColumns],
                    "totals_row_shown": bool(table.totalsRowShown),
                    "style": table.tableStyleInfo.name if table.tableStyleInfo else None,
                }
                for table in source_sheet.tables.values()
            ]
            imported_sheet.metadata["data_validations"] = [
                {
                    "sqref": str(validation.sqref),
                    "type": validation.type,
                    "operator": validation.operator,
                    "formula1": validation.formula1,
                    "formula2": validation.formula2,
                    "allow_blank": validation.allow_blank,
                    "show_error_message": validation.showErrorMessage,
                    "error_title": validation.errorTitle,
                    "error": validation.error,
                }
                for validation in source_sheet.data_validations.dataValidation
            ]
            imported_sheet.metadata["conditional_formats"] = self._extract_conditional_formats(source_sheet)
            value_sheet = values_wb.worksheets[sheet_index] if sheet_index < len(values_wb.worksheets) else None

            max_row = source_sheet.max_row or 0
            max_col = source_sheet.max_column or 0
            imported_addresses: list[str] = []

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    excel_cell = source_sheet.cell(row=row, column=col)
                    value_cell = value_sheet.cell(row=row, column=col) if value_sheet is not None else None
                    address = f"{get_column_letter(col)}{row}"

                    formatting = self._extract_formatting(excel_cell)
                    has_content = excel_cell.value is not None
                    if not has_content and not formatting:
                        continue

                    model_cell = imported_sheet.get_cell(address)
                    imported_addresses.append(address)
                    if isinstance(excel_cell.value, str) and excel_cell.value.startswith("="):
                        model_cell.formula = excel_cell.value
                        model_cell.value = value_cell.value if value_cell is not None else None
                    else:
                        model_cell.value = excel_cell.value
                        model_cell.formula = None

                    if formatting:
                        model_cell.formatting = formatting

            imported_sheet.metadata["ooxml_imported_addresses"] = imported_addresses

        if not workbook.sheets:
            workbook.add_sheet("Sheet1")
        workbook.active_sheet_index = 0
        return workbook

    def export_xlsx(self, path: str, workbook: Workbook) -> None:
        """Export app workbook model data to `.xlsx`."""
        file_path = Path(path)
        try:
            target = self.ooxml_layer.open_template(workbook.metadata)
            using_template = target is not None
            if target is None:
                target = OpenPyxlWorkbook()
                default_sheet = target.active
                target.remove(default_sheet)

            existing_names: set[str] = set()
            existing_table_names: set[str] = set()
            exported_sheet_names: dict[str, str] = {}
            retained_sheets = {
                str(sheet.metadata.get("ooxml_original_title") or sheet.name)
                for sheet in workbook.sheets
            }
            if using_template:
                for existing_sheet in list(target.worksheets):
                    if existing_sheet.title not in retained_sheets:
                        target.remove(existing_sheet)
            for sheet in workbook.sheets:
                sheet_name = self._make_unique_sheet_name(sheet.name or "Sheet", existing_names)
                existing_names.add(sheet_name)
                exported_sheet_names[sheet.name] = sheet_name
                original_title = str(sheet.metadata.get("ooxml_original_title") or "")
                if using_template and original_title in target.sheetnames:
                    target_sheet = target[original_title]
                    target_sheet.title = sheet_name
                    for address in sheet.metadata.get("ooxml_imported_addresses", []):
                        if address not in sheet.cells:
                            target_sheet[str(address)].value = None
                else:
                    target_sheet = target.create_sheet(title=sheet_name)

                freeze_panes = sheet.metadata.get("freeze_panes")
                if freeze_panes:
                    target_sheet.freeze_panes = freeze_panes
                for merged_range in list(target_sheet.merged_cells.ranges):
                    target_sheet.unmerge_cells(str(merged_range))
                for merged_range in sheet.metadata.get("merged_ranges", []):
                    target_sheet.merge_cells(str(merged_range))
                if sheet.metadata.get("auto_filter"):
                    target_sheet.auto_filter.ref = str(sheet.metadata["auto_filter"])
                for column, width in sheet.metadata.get("column_widths", {}).items():
                    target_sheet.column_dimensions[str(column)].width = float(width)
                for row, height in sheet.metadata.get("row_heights", {}).items():
                    target_sheet.row_dimensions[int(row)].height = float(height)

                for address, cell in sheet.cells.items():
                    target_cell = target_sheet[address]
                    if cell.formula:
                        target_cell.value = cell.formula
                    else:
                        target_cell.value = cell.value
                    self._apply_formatting(target_cell, cell.formatting)

                for existing_table in list(target_sheet.tables):
                    del target_sheet.tables[existing_table]
                for table_data in sheet.metadata.get("tables", []):
                    if not isinstance(table_data, dict) or not table_data.get("ref"):
                        continue
                    table_name = self._safe_table_name(
                        str(table_data.get("display_name") or table_data.get("name") or "Table"),
                        existing_table_names,
                    )
                    existing_table_names.add(table_name)
                    table = Table(displayName=table_name, ref=str(table_data["ref"]))
                    style_name = table_data.get("style") or "TableStyleMedium2"
                    table.tableStyleInfo = TableStyleInfo(
                        name=str(style_name), showFirstColumn=False, showLastColumn=False,
                        showRowStripes=True, showColumnStripes=False,
                    )
                    table.totalsRowShown = bool(table_data.get("totals_row_shown", False))
                    target_sheet.add_table(table)

                target_sheet.data_validations.dataValidation = []
                for validation_data in sheet.metadata.get("data_validations", []):
                    if not isinstance(validation_data, dict) or not validation_data.get("sqref"):
                        continue
                    validation = DataValidation(
                        type=validation_data.get("type"),
                        operator=validation_data.get("operator"),
                        formula1=validation_data.get("formula1"),
                        formula2=validation_data.get("formula2"),
                        allow_blank=validation_data.get("allow_blank", False),
                        showErrorMessage=validation_data.get("show_error_message", False),
                        errorTitle=validation_data.get("error_title"),
                        error=validation_data.get("error"),
                    )
                    target_sheet.add_data_validation(validation)
                    validation.add(str(validation_data["sqref"]))

                if not using_template:
                    target_sheet.conditional_formatting = ConditionalFormattingList()
                for rule_data in sheet.metadata.get("conditional_formats", []):
                    if using_template and isinstance(rule_data, dict) and rule_data.get("source") == "imported":
                        continue
                    self._add_conditional_format(target_sheet, rule_data)

                for chart_data in sheet.metadata.get("charts", []):
                    self._add_chart(target_sheet, chart_data)

            if not workbook.sheets:
                target.create_sheet(title="Sheet1")

            if using_template:
                for existing_name, existing_definition in list(target.defined_names.items()):
                    if not existing_definition.hidden:
                        target.defined_names.pop(existing_name, None)
                for target_sheet in target.worksheets:
                    for existing_name,existing_definition in list(target_sheet.defined_names.items()):
                        if not existing_definition.hidden:target_sheet.defined_names.pop(existing_name,None)
            else:
                target.defined_names.clear()
            for item in workbook.metadata.get("defined_names", []):
                if not isinstance(item, dict) or not item.get("name") or not item.get("refers_to"):
                    continue
                scope = item.get("scope")
                exported_scope = exported_sheet_names.get(str(scope),str(scope)) if scope else None
                local_id = target.sheetnames.index(exported_scope) if exported_scope in target.sheetnames else None
                target.defined_names.add(DefinedName(
                    str(item["name"]),
                    attr_text=self._rewrite_defined_name_reference(str(item["refers_to"]).lstrip("="),exported_sheet_names),
                    localSheetId=local_id,
                ))

            target.save(file_path)
        except Exception as exc:  # pragma: no cover - defensive wrapper
            raise WorkbookConversionError(f"Failed to write Excel file: {path}") from exc

    def import_csv(self, path: str) -> Workbook:
        """Import a `.csv` file as a single-sheet workbook."""
        file_path = Path(path)
        workbook = Workbook(name=file_path.stem or "Imported CSV")
        workbook.sheets = []
        sheet = workbook.add_sheet(file_path.stem[:31] or "Sheet1")

        try:
            with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                for row_index, row_values in enumerate(reader, start=1):
                    for col_index, raw_value in enumerate(row_values, start=1):
                        if raw_value == "":
                            continue
                        address = f"{get_column_letter(col_index)}{row_index}"
                        cell = sheet.get_cell(address)
                        if raw_value.startswith("="):
                            cell.formula = raw_value
                            cell.value = None
                        else:
                            cell.value = self._infer_scalar(raw_value)
        except Exception as exc:  # pragma: no cover - defensive wrapper
            raise WorkbookConversionError(f"Failed to read CSV file: {path}") from exc

        return workbook

    def export_csv(self, path: str, workbook: Workbook, sheet_index: int = 0) -> None:
        """Export a worksheet to `.csv`.

        CSV is single-sheet only; by default the active/first sheet is exported.
        """
        if not workbook.sheets:
            raise WorkbookConversionError("Cannot export CSV from an empty workbook")

        normalized_index = max(0, min(sheet_index, len(workbook.sheets) - 1))
        sheet = workbook.sheets[normalized_index]
        if not sheet.cells:
            Path(path).write_text("", encoding="utf-8")
            return

        max_row = 0
        max_col = 0
        for address in sheet.cells:
            row, col = self._address_to_row_col(address)
            max_row = max(max_row, row)
            max_col = max(max_col, col)

        try:
            with Path(path).open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                for row_index in range(1, max_row + 1):
                    row_values: list[Any] = []
                    for col_index in range(1, max_col + 1):
                        address = f"{get_column_letter(col_index)}{row_index}"
                        cell = sheet.cells.get(address)
                        if cell is None:
                            row_values.append("")
                        elif cell.formula:
                            row_values.append(cell.formula)
                        elif cell.value is None:
                            row_values.append("")
                        else:
                            row_values.append(cell.value)
                    writer.writerow(row_values)
        except Exception as exc:  # pragma: no cover - defensive wrapper
            raise WorkbookConversionError(f"Failed to write CSV file: {path}") from exc

    @staticmethod
    def _extract_formatting(excel_cell) -> dict[str, Any]:
        formatting: dict[str, Any] = {}

        if excel_cell.number_format and excel_cell.number_format != "General":
            formatting["number_format"] = excel_cell.number_format

        font = excel_cell.font
        if font is not None:
            if font.bold:
                formatting["bold"] = True
            if font.italic:
                formatting["italic"] = True
            if font.underline and font.underline != "none":
                formatting["underline"] = True
            color = getattr(font, "color", None)
            rgb = getattr(color, "rgb", None)
            if isinstance(rgb, str) and rgb:
                formatting["font_color"] = rgb

        fill = excel_cell.fill
        if fill is not None and getattr(fill, "fill_type", None) == "solid":
            fg_color = getattr(fill, "fgColor", None)
            fg_rgb = getattr(fg_color, "rgb", None)
            if isinstance(fg_rgb, str) and fg_rgb:
                formatting["fill_color"] = fg_rgb

        alignment = excel_cell.alignment
        if alignment is not None:
            if alignment.horizontal:
                formatting["horizontal_align"] = alignment.horizontal
            if alignment.vertical:
                formatting["vertical_align"] = alignment.vertical
            if alignment.wrap_text:
                formatting["wrap_text"] = True

        return formatting

    @staticmethod
    def _extract_conditional_formats(source_sheet) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        for conditional in source_sheet.conditional_formatting:
            for rule in conditional.rules:
                data: dict[str, Any] = {
                    "range": str(conditional.sqref), "type": rule.type,
                    "operator": rule.operator, "formula": list(rule.formula or []),
                    "source": "imported",
                }
                dxf = rule.dxf
                if dxf and dxf.fill and dxf.fill.fill_type == "solid":
                    data["fill_color"] = getattr(dxf.fill.fgColor, "rgb", None)
                if dxf and dxf.font:
                    data["font_color"] = getattr(dxf.font.color, "rgb", None) if dxf.font.color else None
                    data["bold"] = bool(dxf.font.bold)
                result.append(data)
        return result

    @staticmethod
    def _add_conditional_format(target_sheet, data: dict[str, Any]) -> None:
        if not isinstance(data, dict) or not data.get("range"):
            return
        fill_color = data.get("fill_color")
        font_color = data.get("font_color")
        fill = PatternFill(fill_type="solid", fgColor=str(fill_color)) if fill_color else None
        font = Font(color=str(font_color) if font_color else None, bold=bool(data.get("bold")))
        formulas = [str(item) for item in data.get("formula", [])]
        if data.get("type") == "cellIs" and data.get("operator") and formulas:
            rule = CellIsRule(operator=str(data["operator"]), formula=formulas, fill=fill, font=font)
        elif formulas:
            rule = FormulaRule(formula=formulas, fill=fill, font=font)
        else:
            return
        target_sheet.conditional_formatting.add(str(data["range"]), rule)

    @staticmethod
    def _add_chart(target_sheet, data: dict[str, Any]) -> None:
        if not isinstance(data, dict) or not data.get("range"):
            return
        min_col,min_row,max_col,max_row=range_boundaries(str(data["range"]))
        if max_row<=min_row or max_col<min_col:
            return
        chart_type=str(data.get("type") or "column").lower()
        chart={"column":BarChart,"bar":BarChart,"line":LineChart,"pie":PieChart}.get(chart_type,BarChart)()
        chart.title=str(data.get("title") or "Chart")
        if chart_type=="pie":
            data_col=min(min_col+1,max_col)
            chart.add_data(Reference(target_sheet,min_col=data_col,min_row=min_row,max_row=max_row),titles_from_data=True)
        else:
            chart.add_data(Reference(target_sheet,min_col=min_col+1,min_row=min_row,max_col=max_col,max_row=max_row),titles_from_data=True)
        chart.set_categories(Reference(target_sheet,min_col=min_col,min_row=min_row+1,max_row=max_row))
        anchor=str(data.get("anchor") or f"{get_column_letter(max_col+2)}{min_row}")
        for existing in list(target_sheet._charts):
            marker=getattr(existing,"anchor",None)
            if hasattr(marker,"_from") and marker._from.col+1==range_boundaries(f"{anchor}:{anchor}")[0] and marker._from.row+1==range_boundaries(f"{anchor}:{anchor}")[1]:
                target_sheet._charts.remove(existing)
        target_sheet.add_chart(chart,anchor)

    @staticmethod
    def _rewrite_defined_name_reference(reference: str, sheet_names: dict[str,str]) -> str:
        result=reference
        for original,exported in sheet_names.items():
            old_quoted="'"+original.replace("'","''")+"'!"
            new_quoted="'"+exported.replace("'","''")+"'!"
            result=result.replace(old_quoted,new_quoted)
            if " " not in original:result=result.replace(original+"!",new_quoted if " " in exported else exported+"!")
        return result

    @staticmethod
    def _apply_formatting(target_cell, formatting: dict[str, Any]) -> None:
        if not formatting:
            return

        target_cell.number_format = str(formatting.get("number_format", target_cell.number_format))

        has_font_updates = any(
            key in formatting for key in ("bold", "italic", "underline", "font_color")
        )
        if has_font_updates:
            color_value = formatting.get("font_color")
            target_cell.font = Font(
                bold=bool(formatting.get("bold", False)),
                italic=bool(formatting.get("italic", False)),
                underline="single" if formatting.get("underline") else None,
                color=color_value if isinstance(color_value, str) else None,
            )

        fill_color = formatting.get("fill_color")
        if isinstance(fill_color, str) and fill_color:
            target_cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

        has_alignment_updates = any(
            key in formatting for key in ("horizontal_align", "vertical_align", "wrap_text")
        )
        if has_alignment_updates:
            target_cell.alignment = Alignment(
                horizontal=formatting.get("horizontal_align"),
                vertical=formatting.get("vertical_align"),
                wrap_text=bool(formatting.get("wrap_text", False)),
            )

    @staticmethod
    def _address_to_row_col(address: str) -> tuple[int, int]:
        letters = ""
        digits = ""
        for char in address.upper():
            if char.isalpha():
                letters += char
            elif char.isdigit():
                digits += char

        row = int(digits) if digits else 1
        col = 0
        for char in letters:
            col = col * 26 + (ord(char) - ord("A") + 1)
        return row, max(1, col)

    @staticmethod
    def _infer_scalar(raw_value: str) -> Any:
        text = raw_value.strip()
        if text == "":
            return ""

        lower = text.lower()
        if lower == "true":
            return True
        if lower == "false":
            return False

        try:
            if text.isdigit() or (text.startswith("-") and text[1:].isdigit()):
                return int(text)
            return float(text)
        except ValueError:
            return raw_value

    @staticmethod
    def _make_unique_sheet_name(candidate: str, existing_names: set[str]) -> str:
        sanitized = "".join("_" if char in _INVALID_SHEET_CHARS else char for char in candidate).strip()
        if not sanitized:
            sanitized = "Sheet"
        base = sanitized[:31]
        if base not in existing_names:
            return base

        suffix = 2
        while True:
            suffix_text = f" ({suffix})"
            trimmed = base[: max(1, 31 - len(suffix_text))]
            attempt = f"{trimmed}{suffix_text}"
            if attempt not in existing_names:
                return attempt
            suffix += 1

    @staticmethod
    def _safe_table_name(candidate: str, workbook_tables) -> str:
        normalized = "".join(char if char.isalnum() or char == "_" else "_" for char in candidate)
        if not normalized or normalized[0].isdigit():
            normalized = f"Table_{normalized}"
        existing = {name.lower() for name in workbook_tables}
        attempt = normalized
        suffix = 2
        while attempt.lower() in existing:
            attempt = f"{normalized}_{suffix}"
            suffix += 1
        return attempt
