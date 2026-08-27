from openpyxl import load_workbook

from app.models.sheet import Worksheet
from app.models.workbook import Workbook
from app.services.conditional_formatting import formatting_for
from app.services.file_conversion import WorkbookFileConverter


def test_conditional_format_evaluator_applies_matching_style():
    sheet=Worksheet("Data"); sheet.get_cell("A1").value=15
    sheet.metadata["conditional_formats"]=[{"range":"A1:A10","type":"cellIs","operator":"greaterThan","formula":["10"],"fill_color":"00FF00","bold":True}]
    assert formatting_for(sheet,"A1",15)=={"fill_color":"00FF00","bold":True}
    assert formatting_for(sheet,"A2",5)=={}


def test_xlsx_roundtrip_named_ranges_conditional_formats_and_created_chart(tmp_path):
    workbook=Workbook("Compatibility"); sheet=workbook.add_sheet("Data")
    for row,values in enumerate((("Month","Value"),("Jan",10),("Feb",20)),1):
        sheet.get_cell(f"A{row}").value=values[0]; sheet.get_cell(f"B{row}").value=values[1]
    workbook.metadata["defined_names"]=[{"name":"Values","refers_to":"=Data!$B$2:$B$3","scope":None}]
    sheet.metadata["conditional_formats"]=[{"range":"B2:B3","type":"cellIs","operator":"greaterThan","formula":["15"],"fill_color":"00FF00","font_color":"006100"}]
    sheet.metadata["charts"]=[{"type":"column","title":"Monthly values","range":"A1:B3","anchor":"D2"}]
    path=tmp_path/"features.xlsx"; WorkbookFileConverter().export_xlsx(str(path),workbook)
    excel=load_workbook(path)
    assert excel.defined_names["Values"].attr_text=="Data!$B$2:$B$3"
    assert len(excel["Data"].conditional_formatting)==1
    assert len(excel["Data"]._charts)==1
    imported=WorkbookFileConverter().import_xlsx(str(path))
    assert imported.metadata["defined_names"][0]["name"]=="Values"
    assert imported.sheets[0].metadata["conditional_formats"][0]["operator"]=="greaterThan"


def test_defined_names_follow_sanitized_sheet_names(tmp_path):
    workbook=Workbook("Names"); workbook.add_sheet("Bad:/Name").get_cell("A1").value=1
    workbook.metadata["defined_names"]=[{"name":"Target","refers_to":"='Bad:/Name'!$A$1","scope":"Bad:/Name"}]
    path=tmp_path/"names.xlsx"; WorkbookFileConverter().export_xlsx(str(path),workbook)
    excel=load_workbook(path)
    definition=excel["Bad__Name"].defined_names["Target"]
    assert definition.attr_text=="'Bad__Name'!$A$1"
    assert definition.localSheetId==0
    imported=WorkbookFileConverter().import_xlsx(str(path))
    assert imported.metadata["defined_names"][0]["scope"]=="Bad__Name"
