"""Tests for Excel/CSV workbook import-export conversion."""

from __future__ import annotations

import pytest
from openpyxl import Workbook as OpenPyxlWorkbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table
from openpyxl.chart import BarChart, Reference

from app.models.workbook import Workbook
from app.services.file_conversion import WorkbookFileConverter
from app.services.ooxml_preservation import OOXMLPreservationError, OOXMLPreservationLayer


def test_xlsx_roundtrip_preserves_sheet_names_values_formulas_and_core_formatting(tmp_path) -> None:
    converter = WorkbookFileConverter()

    workbook = Workbook(name="Roundtrip")
    workbook.sheets = []

    summary = workbook.add_sheet("Summary")
    data = workbook.add_sheet("Data")

    a1 = summary.get_cell("A1")
    a1.value = 42
    a1.formatting = {
        "number_format": "0.00",
        "bold": True,
        "horizontal_align": "center",
    }

    b1 = summary.get_cell("B1")
    b1.formula = "=A1*2"
    b1.value = 84
    b1.formatting = {"italic": True}

    data.get_cell("A1").value = "hello"

    xlsx_path = tmp_path / "roundtrip.xlsx"
    converter.export_xlsx(str(xlsx_path), workbook)

    loaded = converter.import_xlsx(str(xlsx_path))

    assert [sheet.name for sheet in loaded.sheets] == ["Summary", "Data"]
    assert loaded.sheets[0].cells["A1"].value == 42
    assert loaded.sheets[0].cells["B1"].formula == "=A1*2"
    assert loaded.sheets[0].cells["A1"].formatting["bold"] is True
    assert loaded.sheets[0].cells["A1"].formatting["number_format"] == "0.00"


def test_export_xlsx_sanitizes_duplicate_and_invalid_sheet_names(tmp_path) -> None:
    converter = WorkbookFileConverter()
    workbook = Workbook(name="Sheets")
    workbook.sheets = []

    workbook.add_sheet("Bad:/Name")
    workbook.add_sheet("Bad:/Name")

    path = tmp_path / "names.xlsx"
    converter.export_xlsx(str(path), workbook)

    exported = load_workbook(filename=path)
    assert exported.sheetnames[0] == "Bad__Name"
    assert exported.sheetnames[1] == "Bad__Name (2)"


def test_csv_roundtrip_preserves_formulas_and_scalar_values(tmp_path) -> None:
    converter = WorkbookFileConverter()
    workbook = Workbook(name="CSV Book")
    workbook.sheets = []
    sheet = workbook.add_sheet("Sheet1")

    sheet.get_cell("A1").value = "name"
    sheet.get_cell("B1").value = 10
    sheet.get_cell("C1").formula = "=B1*2"
    sheet.get_cell("A2").value = True

    path = tmp_path / "sheet.csv"
    converter.export_csv(str(path), workbook)

    loaded = converter.import_csv(str(path))
    loaded_sheet = loaded.sheets[0]

    assert loaded_sheet.cells["A1"].value == "name"
    assert loaded_sheet.cells["B1"].value == 10
    assert loaded_sheet.cells["C1"].formula == "=B1*2"
    assert loaded_sheet.cells["A2"].value is True


def test_xlsx_roundtrip_preserves_tables_and_data_validation(tmp_path) -> None:
    source = OpenPyxlWorkbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Category", "Amount"])
    sheet.append(["Hardware", 10])
    sheet.append(["Software", 20])
    sheet.add_table(Table(displayName="SalesTable", ref="A1:B3"))
    validation = DataValidation(type="list", formula1='"Open,Closed"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add("C2:C20")
    source_path = tmp_path / "source.xlsx"
    source.save(source_path)

    converter = WorkbookFileConverter()
    imported = converter.import_xlsx(str(source_path))
    assert imported.sheets[0].metadata["tables"][0]["display_name"] == "SalesTable"
    assert imported.sheets[0].metadata["data_validations"][0]["sqref"] == "C2:C20"

    output_path = tmp_path / "output.xlsx"
    converter.export_xlsx(str(output_path), imported)
    exported = load_workbook(output_path)
    assert "SalesTable" in exported["Data"].tables
    rules = exported["Data"].data_validations.dataValidation
    assert len(rules) == 1
    assert str(rules[0].sqref) == "C2:C20"


def test_ooxml_template_preserves_chart_and_tracks_package_integrity(tmp_path) -> None:
    source = OpenPyxlWorkbook()
    sheet = source.active
    sheet.title = "Metrics"
    sheet.append(["Month", "Value"])
    sheet.append(["Jan", 10])
    sheet.append(["Feb", 20])
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    sheet.add_chart(chart, "D2")
    source_path = tmp_path / "chart.xlsx"
    source.save(source_path)

    converter = WorkbookFileConverter()
    imported = converter.import_xlsx(str(source_path))
    snapshot = imported.metadata["ooxml_passthrough"]
    assert snapshot["size"] > 0
    assert snapshot["preserved_parts"] >= 2
    imported.sheets[0].get_cell("B2").value = 15

    output_path = tmp_path / "chart-output.xlsx"
    converter.export_xlsx(str(output_path), imported)
    exported = load_workbook(output_path)
    assert exported["Metrics"]["B2"].value == 15
    assert len(exported["Metrics"]._charts) == 1

    snapshot["sha256"] = "0" * 64
    with pytest.raises(OOXMLPreservationError):
        OOXMLPreservationLayer().restore_bytes(imported.metadata)
